"""Write policy data to Excel files (testing mode) or Axis CRM API (production)."""

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from pathlib import Path

log = logging.getLogger(__name__)

OUTPUT_DIR = Path(os.getenv("POLICY_OUTPUT_DIR", "output"))


def upsert_policies(adviser_id: str, portal_id: str, policies: list[dict]) -> str:
    """Write policies to an Excel file, one sheet per sync.

    Returns the path to the generated file.
    """
    import openpyxl

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = f"{portal_id}_{adviser_id}_{ts}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{portal_id} policies"

    # Header row
    headers = [
        "Policy Number", "Client Name", "Product Name", "Status",
        "Premium Amount", "Premium Frequency", "Sum Insured",
        "Policy Start Date", "Next Payment Date",
    ]
    ws.append(headers)

    # Bold headers
    from openpyxl.styles import Font
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    # Data rows
    for p in policies:
        ws.append([
            p.get("policy_number", ""),
            p.get("client_name", ""),
            p.get("product_name", ""),
            p.get("status", ""),
            p.get("premium_amount", ""),
            p.get("premium_frequency", ""),
            p.get("sum_insured", ""),
            p.get("policy_start_date", ""),
            p.get("next_payment_date", ""),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Summary sheet
    ws2 = wb.create_sheet("Sync Info")
    ws2.append(["Field", "Value"])
    ws2.append(["Adviser ID", adviser_id])
    ws2.append(["Portal", portal_id])
    ws2.append(["Synced At", datetime.now(timezone.utc).isoformat()])
    ws2.append(["Total Policies", len(policies)])

    wb.save(filepath)
    log.info("Wrote %d policies to %s", len(policies), filepath)
    return str(filepath)
